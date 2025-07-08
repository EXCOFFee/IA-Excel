#!/usr/bin/env python3
"""
Creador de Instalador para Planificador Inteligente
==================================================

Este script crea un instalador completo para el Planificador Inteligente
"""

import os
import shutil
import zipfile
from pathlib import Path
import json
from datetime import datetime

def crear_instalador():
    """Crea el paquete instalador completo"""
    
    print("🚀 Creando instalador del Planificador Inteligente...")
    
    # Configuración
    version = "1.0.0"
    nombre_producto = "Planificador_Inteligente"
    fecha_build = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Directorios
    base_dir = Path(__file__).parent
    dist_dir = base_dir / "dist"
    build_dir = base_dir / "installer_build"
    
    # Limpiar build anterior
    if build_dir.exists():
        shutil.rmtree(build_dir)
    
    build_dir.mkdir(parents=True, exist_ok=True)
    
    # Crear estructura del instalador
    app_dir = build_dir / nombre_producto
    app_dir.mkdir(parents=True, exist_ok=True)
    
    print("📦 Copiando archivos...")
    
    # Copiar ejecutable
    exe_source = dist_dir / "PlanificadorInteligente.exe"
    exe_dest = app_dir / "PlanificadorInteligente.exe"
    
    if exe_source.exists():
        shutil.copy2(exe_source, exe_dest)
        print(f"✅ Ejecutable copiado: {exe_dest}")
    else:
        print("❌ Error: No se encontró el ejecutable")
        return False
    
    # Copiar documentación
    docs_a_copiar = [
        "README.md",
        "README_USUARIO.md",
        "GUIA_INTEGRACION_EXCEL.md",
        "ESTADO_PRODUCCION.md",
        "NOTAS_VERSION.md"
    ]
    
    docs_dir = app_dir / "docs"
    docs_dir.mkdir(exist_ok=True)
    
    for doc in docs_a_copiar:
        source = base_dir / doc
        if source.exists():
            shutil.copy2(source, docs_dir / doc)
            print(f"✅ Documentación copiada: {doc}")
    
    # Crear archivo de configuración
    config = {
        "version": version,
        "nombre": "Planificador Inteligente",
        "descripcion": "Sistema de planificación inteligente con procesamiento de Excel",
        "autor": "Sistema de Planificación IA",
        "fecha_build": fecha_build,
        "archivos": {
            "ejecutable": "PlanificadorInteligente.exe",
            "documentacion": "docs/",
            "plantilla_excel": "plantilla_planificacion.xlsx"
        }
    }
    
    config_file = app_dir / "config.json"
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    
    print("✅ Configuración creada")
    
    # Crear plantilla de Excel
    crear_plantilla_excel(app_dir)
    
    # Crear script de instalación
    crear_script_instalacion(app_dir)
    
    # Crear archivo batch para ejecutar
    crear_batch_ejecutor(app_dir)
    
    # Crear archivo de desinstalación
    crear_desinstalador(app_dir)
    
    # Crear archivo README del instalador
    crear_readme_instalador(app_dir)
    
    # Crear archivo ZIP del instalador
    zip_filename = f"{nombre_producto}_v{version}_{fecha_build}.zip"
    zip_path = build_dir / zip_filename
    
    print("📦 Creando archivo ZIP...")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(app_dir):
            for file in files:
                file_path = Path(root) / file
                arcname = file_path.relative_to(app_dir)
                zipf.write(file_path, arcname)
    
    print(f"✅ Instalador creado: {zip_path}")
    print(f"📊 Tamaño: {zip_path.stat().st_size / (1024*1024):.1f} MB")
    
    return True

def crear_plantilla_excel(app_dir):
    """Crea una plantilla de Excel básica"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planificación"
        
        # Encabezados
        headers = [
            "ID_Proceso", "Nombre_Proceso", "Descripcion", "Prioridad",
            "Tiempo_Estimado", "Recursos_Necesarios", "Fecha_Inicio",
            "Fecha_Limite", "Estado", "Observaciones"
        ]
        
        # Estilo de encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos de ejemplo
        datos_ejemplo = [
            ["PROC001", "Análisis de Datos", "Procesamiento inicial de datos", "Alta", 
             "4 horas", "Analista,Servidor", "2024-01-15", "2024-01-20", "Pendiente", "Urgente"],
            ["PROC002", "Generación de Reportes", "Crear reportes mensuales", "Media", 
             "2 horas", "Analista", "2024-01-16", "2024-01-25", "En Proceso", ""],
            ["PROC003", "Validación de Resultados", "Verificar exactitud", "Alta", 
             "3 horas", "Supervisor", "2024-01-18", "2024-01-22", "Pendiente", "Revisar criterios"]
        ]
        
        # Escribir datos de ejemplo
        for row, datos in enumerate(datos_ejemplo, 2):
            for col, dato in enumerate(datos, 1):
                ws.cell(row=row, column=col, value=dato)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        plantilla_path = app_dir / "plantilla_planificacion.xlsx"
        wb.save(plantilla_path)
        
        print("✅ Plantilla Excel creada")
        
    except ImportError:
        print("⚠️ No se pudo crear la plantilla Excel (openpyxl no disponible)")
        # Crear archivo de texto alternativo
        plantilla_path = app_dir / "plantilla_planificacion.txt"
        with open(plantilla_path, 'w', encoding='utf-8') as f:
            f.write("Plantilla de Planificación - Use Excel para mejor experiencia\n")
            f.write("=" * 50 + "\n\n")
            f.write("Columnas requeridas:\n")
            f.write("- ID_Proceso\n")
            f.write("- Nombre_Proceso\n")
            f.write("- Descripcion\n")
            f.write("- Prioridad (Alta/Media/Baja)\n")
            f.write("- Tiempo_Estimado\n")
            f.write("- Recursos_Necesarios\n")
            f.write("- Fecha_Inicio\n")
            f.write("- Fecha_Limite\n")
            f.write("- Estado\n")
            f.write("- Observaciones\n")

def crear_script_instalacion(app_dir):
    """Crea script de instalación para Windows"""
    
    script_content = '''@echo off
echo ====================================
echo   Planificador Inteligente v1.0.0
echo   Instalador Automatico
echo ====================================
echo.

echo Creando directorio de instalacion...
set "INSTALL_DIR=%USERPROFILE%\\PlanificadorInteligente"
if not exist "%INSTALL_DIR%" mkdir "%INSTALL_DIR%"

echo Copiando archivos...
copy /Y "PlanificadorInteligente.exe" "%INSTALL_DIR%\\"
copy /Y "config.json" "%INSTALL_DIR%\\"
copy /Y "plantilla_planificacion.*" "%INSTALL_DIR%\\"

echo Copiando documentacion...
if not exist "%INSTALL_DIR%\\docs" mkdir "%INSTALL_DIR%\\docs"
copy /Y "docs\\*" "%INSTALL_DIR%\\docs\\"

echo Creando acceso directo en el escritorio...
set "DESKTOP=%USERPROFILE%\\Desktop"
echo Set WshShell = WScript.CreateObject("WScript.Shell") > "%TEMP%\\shortcut.vbs"
echo Set Shortcut = WshShell.CreateShortcut("%DESKTOP%\\Planificador Inteligente.lnk") >> "%TEMP%\\shortcut.vbs"
echo Shortcut.TargetPath = "%INSTALL_DIR%\\PlanificadorInteligente.exe" >> "%TEMP%\\shortcut.vbs"
echo Shortcut.WorkingDirectory = "%INSTALL_DIR%" >> "%TEMP%\\shortcut.vbs"
echo Shortcut.Description = "Planificador Inteligente - Sistema de Planificacion" >> "%TEMP%\\shortcut.vbs"
echo Shortcut.Save >> "%TEMP%\\shortcut.vbs"
cscript //nologo "%TEMP%\\shortcut.vbs"
del "%TEMP%\\shortcut.vbs"

echo.
echo ====================================
echo   Instalacion completada!
echo ====================================
echo.
echo El programa se ha instalado en: %INSTALL_DIR%
echo Se ha creado un acceso directo en el escritorio.
echo.
echo Para ejecutar el programa:
echo - Haga doble clic en el acceso directo del escritorio
echo - O navegue a %INSTALL_DIR% y ejecute PlanificadorInteligente.exe
echo.
pause
'''
    
    script_path = app_dir / "instalar.bat"
    with open(script_path, 'w', encoding='utf-8') as f:
        f.write(script_content)
    
    print("✅ Script de instalación creado")

def crear_batch_ejecutor(app_dir):
    """Crea archivo batch para ejecutar fácilmente"""
    
    batch_content = '''@echo off
title Planificador Inteligente
echo Iniciando Planificador Inteligente...
echo.
start "" "PlanificadorInteligente.exe"
'''
    
    batch_path = app_dir / "Ejecutar_Planificador.bat"
    with open(batch_path, 'w', encoding='utf-8') as f:
        f.write(batch_content)
    
    print("✅ Ejecutor batch creado")

def crear_desinstalador(app_dir):
    """Crea script de desinstalación"""
    
    uninstall_content = '''@echo off
echo ====================================
echo   Planificador Inteligente v1.0.0
echo   Desinstalador
echo ====================================
echo.

set "INSTALL_DIR=%USERPROFILE%\\PlanificadorInteligente"
set "DESKTOP=%USERPROFILE%\\Desktop"

echo ¿Está seguro que desea desinstalar el Planificador Inteligente?
pause

echo Eliminando archivos...
if exist "%INSTALL_DIR%" (
    rmdir /S /Q "%INSTALL_DIR%"
    echo Archivos eliminados.
) else (
    echo No se encontraron archivos de instalación.
)

echo Eliminando acceso directo...
if exist "%DESKTOP%\\Planificador Inteligente.lnk" (
    del "%DESKTOP%\\Planificador Inteligente.lnk"
    echo Acceso directo eliminado.
)

echo.
echo ====================================
echo   Desinstalación completada
echo ====================================
echo.
pause
'''
    
    uninstall_path = app_dir / "desinstalar.bat"
    with open(uninstall_path, 'w', encoding='utf-8') as f:
        f.write(uninstall_content)
    
    print("✅ Desinstalador creado")

def crear_readme_instalador(app_dir):
    """Crea README para el instalador"""
    
    readme_content = '''# 🚀 Planificador Inteligente - Instalador

## 📋 Contenido del Paquete

Este paquete contiene todo lo necesario para instalar y usar el Planificador Inteligente:

### 📁 Archivos Incluidos
- `PlanificadorInteligente.exe` - Aplicación principal
- `instalar.bat` - Script de instalación automática
- `Ejecutar_Planificador.bat` - Ejecutor directo
- `desinstalar.bat` - Script de desinstalación
- `config.json` - Configuración del sistema
- `plantilla_planificacion.xlsx` - Plantilla de Excel
- `docs/` - Documentación completa

## 🔧 Instalación

### Opción 1: Instalación Automática (Recomendada)
1. Extraer todos los archivos a una carpeta
2. Ejecutar `instalar.bat` como administrador
3. Seguir las instrucciones en pantalla
4. Usar el acceso directo creado en el escritorio

### Opción 2: Instalación Manual
1. Extraer archivos a la carpeta deseada
2. Ejecutar `PlanificadorInteligente.exe` directamente
3. Crear acceso directo manualmente si es necesario

## 📊 Uso del Sistema

### Primeros Pasos
1. Abrir el Planificador Inteligente
2. Descargar la plantilla de Excel
3. Llenar los datos en Excel
4. Cargar el archivo en el sistema
5. Procesar y obtener resultados

### Funcionalidades Principales
- ✅ Procesamiento de archivos Excel
- ✅ Análisis inteligente de procesos
- ✅ Optimización de recursos
- ✅ Generación de reportes
- ✅ Interfaz gráfica amigable
- ✅ Exportación de resultados

## 🛠️ Requisitos del Sistema

- **Sistema Operativo**: Windows 10 o superior
- **Memoria RAM**: 4 GB mínimo (8 GB recomendado)
- **Espacio en Disco**: 500 MB disponibles
- **Software**: Microsoft Excel (opcional, para mejor experiencia)

## 📞 Soporte

### Documentación
- Ver carpeta `docs/` para documentación completa
- `README_USUARIO.md` - Guía de usuario
- `GUIA_INTEGRACION_EXCEL.md` - Guía de Excel

### Resolución de Problemas
1. Verificar que Windows Defender no bloquee la aplicación
2. Ejecutar como administrador si es necesario
3. Verificar que los archivos estén completos
4. Consultar los logs de error en caso de problemas

## 🔄 Desinstalación

Para desinstalar el programa:
1. Ejecutar `desinstalar.bat`
2. Confirmar la desinstalación
3. Los archivos y accesos directos serán eliminados

## 📈 Versión

**Versión**: 1.0.0
**Fecha**: 2024
**Estado**: Producción

## 🏆 Características

- **Multiplataforma**: Funciona en Windows
- **Sin dependencias**: No requiere Python instalado
- **Interfaz intuitiva**: Fácil de usar
- **Procesamiento rápido**: Optimizado para rendimiento
- **Documentación completa**: Guías y manuales incluidos

¡Gracias por usar el Planificador Inteligente! 🚀
'''
    
    readme_path = app_dir / "LEEME_INSTALADOR.md"
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print("✅ README del instalador creado")

if __name__ == "__main__":
    try:
        if crear_instalador():
            print("\n🎉 ¡Instalador creado exitosamente!")
            print("El archivo ZIP está listo para distribuir.")
        else:
            print("\n❌ Error al crear el instalador.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        input("Presione Enter para continuar...")
