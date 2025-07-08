"""
Lector de Archivos Excel

Este módulo implementa la funcionalidad para leer y procesar archivos Excel,
extrayendo datos de procesos y recursos para convertirlos en entidades
del dominio del sistema.

Funcionalidades:
- Lectura de archivos Excel (.xlsx, .xls)
- Validación de formato y estructura
- Conversión a entidades del dominio
- Manejo de errores y excepciones
- Soporte para múltiples hojas de cálculo

Principios SOLID aplicados:
- Single Responsibility: Solo se encarga de leer archivos Excel
- Open/Closed: Extensible para nuevos formatos
- Dependency Inversion: No depende de implementaciones concretas

Autor: Equipo de Desarrollo
Fecha: 2025-07-07
"""

from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging
from datetime import datetime
import re

from domain.models.proceso import Proceso, TipoProceso, EstadoProceso, NivelPrioridad
from domain.models.recurso import Recurso, TipoRecurso, EstadoRecurso, NivelExperiencia, HorarioTrabajo
from datetime import time


# Configuración de logging
logger = logging.getLogger(__name__)


@dataclass
class ConfiguracionLectura:
    """
    Configuración para la lectura de archivos Excel.
    
    Attributes:
        hoja_procesos: Nombre de la hoja con datos de procesos
        hoja_recursos: Nombre de la hoja con datos de recursos
        fila_inicio: Fila donde inician los datos (1-indexed)
        validar_formato: Si se debe validar el formato estrictamente
        permitir_campos_vacios: Si se permiten campos vacíos
        encoding: Codificación del archivo
    """
    hoja_procesos: str = "Procesos"
    hoja_recursos: str = "Recursos"
    fila_inicio: int = 2  # Asume que fila 1 tiene headers
    validar_formato: bool = True
    permitir_campos_vacios: bool = False
    encoding: str = "utf-8"


@dataclass
class ResultadoLectura:
    """
    Resultado de la lectura de un archivo Excel.
    
    Attributes:
        procesos: Lista de procesos leídos
        recursos: Lista de recursos leídos
        errores: Lista de errores encontrados
        advertencias: Lista de advertencias
        estadisticas: Estadísticas de la lectura
    """
    procesos: List[Proceso]
    recursos: List[Recurso]
    errores: List[str]
    advertencias: List[str]
    estadisticas: Dict[str, Any]


class LectorExcel:
    """
    Clase para leer archivos Excel y convertir datos a entidades del dominio.
    
    Esta clase implementa la lógica necesaria para procesar archivos Excel
    que contienen información de procesos y recursos, validando los datos
    y convirtiéndolos en entidades del dominio.
    """
    
    def __init__(self, configuracion: Optional[ConfiguracionLectura] = None):
        """
        Inicializa el lector Excel.
        
        Args:
            configuracion: Configuración para la lectura
        """
        self._configuracion = configuracion or ConfiguracionLectura()
        self._logger = logging.getLogger(self.__class__.__name__)
        
        # Mapeo de columnas esperadas
        self._columnas_procesos = {
            'nombre': ['nombre', 'name', 'proceso', 'process'],
            'descripcion': ['descripcion', 'description', 'desc'],
            'tipo': ['tipo', 'type', 'categoria', 'category'],
            'tiempo_estimado': ['tiempo_estimado', 'tiempo', 'horas', 'duracion', 'duration'],
            'prioridad': ['prioridad', 'priority', 'nivel_prioridad'],
            'recursos_requeridos': ['recursos', 'resources', 'recursos_requeridos'],
            'fecha_limite': ['fecha_limite', 'deadline', 'vencimiento'],
            'asignado_a': ['asignado_a', 'assigned_to', 'responsable'],
            'notas': ['notas', 'notes', 'observaciones', 'comments']
        }
        
        self._columnas_recursos = {
            'nombre': ['nombre', 'name', 'recurso', 'resource'],
            'tipo': ['tipo', 'type', 'categoria', 'category'],
            'capacidad_maxima': ['capacidad', 'capacity', 'capacidad_maxima'],
            'costo_por_hora': ['costo', 'cost', 'precio', 'price', 'tarifa'],
            'ubicacion': ['ubicacion', 'location', 'lugar'],
            'responsable': ['responsable', 'manager', 'encargado'],
            'habilidades': ['habilidades', 'skills', 'competencias'],
            'experiencia': ['experiencia', 'experience', 'nivel'],
            'horario_inicio': ['horario_inicio', 'hora_inicio', 'start_time'],
            'horario_fin': ['horario_fin', 'hora_fin', 'end_time']
        }
    
    def leer_archivo(self, ruta_archivo: str) -> ResultadoLectura:
        """
        Lee un archivo Excel y extrae los datos de procesos y recursos.
        
        Args:
            ruta_archivo: Ruta del archivo Excel a leer
            
        Returns:
            ResultadoLectura: Resultado de la lectura con datos procesados
            
        Raises:
            FileNotFoundError: Si el archivo no existe
            ValueError: Si el archivo tiene formato inválido
            RuntimeError: Si ocurre un error durante la lectura
        """
        try:
            self._logger.info(f"Iniciando lectura de archivo: {ruta_archivo}")
            
            # Validar que el archivo existe
            if not Path(ruta_archivo).exists():
                raise FileNotFoundError(f"El archivo {ruta_archivo} no existe")
            
            # Validar extensión
            if not ruta_archivo.lower().endswith(('.xlsx', '.xls')):
                raise ValueError("El archivo debe ser un archivo Excel (.xlsx o .xls)")
            
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
            
            # Inicializar resultado
            resultado = ResultadoLectura(
                procesos=[],
                recursos=[],
                errores=[],
                advertencias=[],
                estadisticas={}
            )
            
            # Leer procesos
            if self._configuracion.hoja_procesos in workbook.sheetnames:
                procesos, errores_procesos = self._leer_procesos(workbook[self._configuracion.hoja_procesos])
                resultado.procesos = procesos
                resultado.errores.extend(errores_procesos)
            else:
                resultado.advertencias.append(f"Hoja '{self._configuracion.hoja_procesos}' no encontrada")
            
            # Leer recursos
            if self._configuracion.hoja_recursos in workbook.sheetnames:
                recursos, errores_recursos = self._leer_recursos(workbook[self._configuracion.hoja_recursos])
                resultado.recursos = recursos
                resultado.errores.extend(errores_recursos)
            else:
                resultado.advertencias.append(f"Hoja '{self._configuracion.hoja_recursos}' no encontrada")
            
            # Calcular estadísticas
            resultado.estadisticas = self._calcular_estadisticas(resultado)
            
            self._logger.info(f"Lectura completada: {len(resultado.procesos)} procesos, {len(resultado.recursos)} recursos")
            
            return resultado
            
        except Exception as e:
            self._logger.error(f"Error leyendo archivo Excel: {str(e)}")
            raise RuntimeError(f"Error leyendo archivo Excel: {str(e)}")
    
    def _leer_procesos(self, hoja: Worksheet) -> Tuple[List[Proceso], List[str]]:
        """
        Lee los datos de procesos desde una hoja de Excel.
        
        Args:
            hoja: Hoja de Excel con datos de procesos
            
        Returns:
            Tuple[List[Proceso], List[str]]: Lista de procesos y errores
        """
        procesos = []
        errores = []
        
        try:
            # Convertir hoja a DataFrame para facilitar el procesamiento
            data = []
            for row in hoja.iter_rows(min_row=1, values_only=True):
                data.append(row)
            
            if not data:
                errores.append("La hoja de procesos está vacía")
                return procesos, errores
            
            # Crear DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])  # Primera fila como headers
            
            # Normalizar nombres de columnas
            df.columns = [str(col).lower().strip() for col in df.columns]
            
            # Mapear columnas
            mapeo_columnas = self._mapear_columnas(df.columns, self._columnas_procesos)
            
            # Procesar cada fila
            for idx, row in df.iterrows():
                try:
                    proceso = self._crear_proceso_desde_fila(row, mapeo_columnas)
                    if proceso:
                        procesos.append(proceso)
                except Exception as e:
                    errores.append(f"Error en fila {idx + 2}: {str(e)}")
            
        except Exception as e:
            errores.append(f"Error procesando hoja de procesos: {str(e)}")
        
        return procesos, errores
    
    def _leer_recursos(self, hoja: Worksheet) -> Tuple[List[Recurso], List[str]]:
        """
        Lee los datos de recursos desde una hoja de Excel.
        
        Args:
            hoja: Hoja de Excel con datos de recursos
            
        Returns:
            Tuple[List[Recurso], List[str]]: Lista de recursos y errores
        """
        recursos = []
        errores = []
        
        try:
            # Convertir hoja a DataFrame
            data = []
            for row in hoja.iter_rows(min_row=1, values_only=True):
                data.append(row)
            
            if not data:
                errores.append("La hoja de recursos está vacía")
                return recursos, errores
            
            # Crear DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])
            
            # Normalizar nombres de columnas
            df.columns = [str(col).lower().strip() for col in df.columns]
            
            # Mapear columnas
            mapeo_columnas = self._mapear_columnas(df.columns, self._columnas_recursos)
            
            # Procesar cada fila
            for idx, row in df.iterrows():
                try:
                    recurso = self._crear_recurso_desde_fila(row, mapeo_columnas)
                    if recurso:
                        recursos.append(recurso)
                except Exception as e:
                    errores.append(f"Error en fila {idx + 2}: {str(e)}")
            
        except Exception as e:
            errores.append(f"Error procesando hoja de recursos: {str(e)}")
        
        return recursos, errores
    
    def _mapear_columnas(self, columnas_archivo: List[str], columnas_esperadas: Dict[str, List[str]]) -> Dict[str, Optional[str]]:
        """
        Mapea las columnas del archivo con las columnas esperadas.
        
        Args:
            columnas_archivo: Lista de nombres de columnas en el archivo
            columnas_esperadas: Diccionario con columnas esperadas
            
        Returns:
            Dict[str, Optional[str]]: Mapeo de columnas
        """
        mapeo = {}
        
        for campo, posibles_nombres in columnas_esperadas.items():
            columna_encontrada = None
            
            for nombre_posible in posibles_nombres:
                if nombre_posible.lower() in columnas_archivo:
                    columna_encontrada = nombre_posible.lower()
                    break
            
            mapeo[campo] = columna_encontrada
        
        return mapeo
    
    def _crear_proceso_desde_fila(self, fila: pd.Series, mapeo: Dict[str, Optional[str]]) -> Optional[Proceso]:
        """
        Crea un proceso desde una fila de datos.
        
        Args:
            fila: Fila de datos del DataFrame
            mapeo: Mapeo de columnas
            
        Returns:
            Optional[Proceso]: Proceso creado o None si hay errores
        """
        try:
            # Extraer datos obligatorios
            nombre = self._extraer_valor(fila, mapeo['nombre'])
            if not nombre:
                return None
            
            descripcion = self._extraer_valor(fila, mapeo['descripcion'], "")
            
            # Tipo de proceso
            tipo_str = self._extraer_valor(fila, mapeo['tipo'], "rutinario")
            tipo = self._convertir_tipo_proceso(tipo_str)
            
            # Tiempo estimado
            tiempo_estimado = self._extraer_valor_numerico(fila, mapeo['tiempo_estimado'])
            if tiempo_estimado is None or tiempo_estimado <= 0:
                raise ValueError("Tiempo estimado inválido")
            
            # Prioridad
            prioridad_str = self._extraer_valor(fila, mapeo['prioridad'], "media")
            prioridad = self._convertir_prioridad(prioridad_str)
            
            # Crear proceso
            proceso = Proceso(
                nombre=nombre,
                descripcion=descripcion,
                tipo=tipo,
                tiempo_estimado_horas=tiempo_estimado,
                prioridad=prioridad
            )
            
            # Campos opcionales
            recursos_req = self._extraer_valor(fila, mapeo['recursos_requeridos'])
            if recursos_req:
                proceso.recursos_requeridos = [r.strip() for r in str(recursos_req).split(',')]
            
            # Fecha límite
            fecha_limite = self._extraer_fecha(fila, mapeo['fecha_limite'])
            if fecha_limite:
                proceso.fecha_limite = fecha_limite
            
            # Asignado a
            asignado_a = self._extraer_valor(fila, mapeo['asignado_a'])
            if asignado_a:
                proceso.asignado_a = str(asignado_a)
            
            # Notas
            notas = self._extraer_valor(fila, mapeo['notas'])
            if notas:
                proceso.notas = str(notas)
            
            return proceso
            
        except Exception as e:
            self._logger.error(f"Error creando proceso: {str(e)}")
            return None
    
    def _crear_recurso_desde_fila(self, fila: pd.Series, mapeo: Dict[str, Optional[str]]) -> Optional[Recurso]:
        """
        Crea un recurso desde una fila de datos.
        
        Args:
            fila: Fila de datos del DataFrame
            mapeo: Mapeo de columnas
            
        Returns:
            Optional[Recurso]: Recurso creado o None si hay errores
        """
        try:
            # Extraer datos obligatorios
            nombre = self._extraer_valor(fila, mapeo['nombre'])
            if not nombre:
                return None
            
            # Tipo de recurso
            tipo_str = self._extraer_valor(fila, mapeo['tipo'], "humano")
            tipo = self._convertir_tipo_recurso(tipo_str)
            
            # Capacidad máxima
            capacidad = self._extraer_valor_numerico(fila, mapeo['capacidad_maxima'])
            if capacidad is None or capacidad <= 0:
                raise ValueError("Capacidad máxima inválida")
            
            # Crear recurso
            recurso = Recurso(
                nombre=nombre,
                tipo=tipo,
                capacidad_maxima=capacidad
            )
            
            # Campos opcionales
            costo = self._extraer_valor_numerico(fila, mapeo['costo_por_hora'])
            if costo is not None:
                recurso.costo_por_hora = costo
            
            ubicacion = self._extraer_valor(fila, mapeo['ubicacion'])
            if ubicacion:
                recurso.ubicacion = str(ubicacion)
            
            responsable = self._extraer_valor(fila, mapeo['responsable'])
            if responsable:
                recurso.responsable = str(responsable)
            
            # Habilidades
            habilidades = self._extraer_valor(fila, mapeo['habilidades'])
            if habilidades:
                recurso.habilidades = [h.strip() for h in str(habilidades).split(',')]
            
            # Experiencia
            experiencia_str = self._extraer_valor(fila, mapeo['experiencia'])
            if experiencia_str:
                recurso.experiencia = self._convertir_experiencia(experiencia_str)
            
            # Horario
            horario = self._crear_horario_desde_fila(fila, mapeo)
            if horario:
                recurso.horario = horario
            
            return recurso
            
        except Exception as e:
            self._logger.error(f"Error creando recurso: {str(e)}")
            return None
    
    def _crear_horario_desde_fila(self, fila: pd.Series, mapeo: Dict[str, Optional[str]]) -> Optional[HorarioTrabajo]:
        """
        Crea un horario de trabajo desde una fila de datos.
        
        Args:
            fila: Fila de datos
            mapeo: Mapeo de columnas
            
        Returns:
            Optional[HorarioTrabajo]: Horario creado o None
        """
        try:
            hora_inicio_str = self._extraer_valor(fila, mapeo['horario_inicio'])
            hora_fin_str = self._extraer_valor(fila, mapeo['horario_fin'])
            
            if hora_inicio_str and hora_fin_str:
                hora_inicio = self._convertir_hora(hora_inicio_str)
                hora_fin = self._convertir_hora(hora_fin_str)
                
                if hora_inicio and hora_fin:
                    return HorarioTrabajo(
                        hora_inicio=hora_inicio,
                        hora_fin=hora_fin
                    )
            
            return None
            
        except Exception:
            return None
    
    def _extraer_valor(self, fila: pd.Series, columna: Optional[str], default: Any = None) -> Any:
        """
        Extrae un valor de una fila.
        
        Args:
            fila: Fila de datos
            columna: Nombre de la columna
            default: Valor por defecto
            
        Returns:
            Any: Valor extraído
        """
        if columna is None or columna not in fila.index:
            return default
        
        valor = fila[columna]
        
        # Manejar valores nulos
        if pd.isna(valor) or valor == "":
            return default
        
        return valor
    
    def _extraer_valor_numerico(self, fila: pd.Series, columna: Optional[str]) -> Optional[float]:
        """
        Extrae un valor numérico de una fila.
        
        Args:
            fila: Fila de datos
            columna: Nombre de la columna
            
        Returns:
            Optional[float]: Valor numérico o None
        """
        valor = self._extraer_valor(fila, columna)
        
        if valor is None:
            return None
        
        try:
            return float(valor)
        except (ValueError, TypeError):
            return None
    
    def _extraer_fecha(self, fila: pd.Series, columna: Optional[str]) -> Optional[datetime]:
        """
        Extrae una fecha de una fila.
        
        Args:
            fila: Fila de datos
            columna: Nombre de la columna
            
        Returns:
            Optional[datetime]: Fecha extraída o None
        """
        valor = self._extraer_valor(fila, columna)
        
        if valor is None:
            return None
        
        try:
            if isinstance(valor, datetime):
                return valor
            
            # Intentar parsear como string
            return pd.to_datetime(valor)
        except (ValueError, TypeError):
            return None
    
    def _convertir_tipo_proceso(self, tipo_str: str) -> TipoProceso:
        """
        Convierte string a TipoProceso.
        
        Args:
            tipo_str: String con el tipo
            
        Returns:
            TipoProceso: Tipo de proceso
        """
        tipo_str = str(tipo_str).lower().strip()
        
        mapeo = {
            'rutinario': TipoProceso.RUTINARIO,
            'especial': TipoProceso.ESPECIAL,
            'urgente': TipoProceso.URGENTE,
            'mantenimiento': TipoProceso.MANTENIMIENTO
        }
        
        return mapeo.get(tipo_str, TipoProceso.RUTINARIO)
    
    def _convertir_prioridad(self, prioridad_str: str) -> NivelPrioridad:
        """
        Convierte string a NivelPrioridad.
        
        Args:
            prioridad_str: String con la prioridad
            
        Returns:
            NivelPrioridad: Nivel de prioridad
        """
        prioridad_str = str(prioridad_str).lower().strip()
        
        mapeo = {
            'baja': NivelPrioridad.BAJA,
            'media': NivelPrioridad.MEDIA,
            'alta': NivelPrioridad.ALTA,
            'critica': NivelPrioridad.CRITICA,
            'crítica': NivelPrioridad.CRITICA
        }
        
        return mapeo.get(prioridad_str, NivelPrioridad.MEDIA)
    
    def _convertir_tipo_recurso(self, tipo_str: str) -> TipoRecurso:
        """
        Convierte string a TipoRecurso.
        
        Args:
            tipo_str: String con el tipo
            
        Returns:
            TipoRecurso: Tipo de recurso
        """
        tipo_str = str(tipo_str).lower().strip()
        
        mapeo = {
            'humano': TipoRecurso.HUMANO,
            'material': TipoRecurso.MATERIAL,
            'tecnologico': TipoRecurso.TECNOLOGICO,
            'tecnológico': TipoRecurso.TECNOLOGICO,
            'espacial': TipoRecurso.ESPACIAL,
            'financiero': TipoRecurso.FINANCIERO
        }
        
        return mapeo.get(tipo_str, TipoRecurso.HUMANO)
    
    def _convertir_experiencia(self, experiencia_str: str) -> NivelExperiencia:
        """
        Convierte string a NivelExperiencia.
        
        Args:
            experiencia_str: String con la experiencia
            
        Returns:
            NivelExperiencia: Nivel de experiencia
        """
        experiencia_str = str(experiencia_str).lower().strip()
        
        mapeo = {
            'junior': NivelExperiencia.JUNIOR,
            'intermedio': NivelExperiencia.INTERMEDIO,
            'senior': NivelExperiencia.SENIOR,
            'experto': NivelExperiencia.EXPERTO
        }
        
        return mapeo.get(experiencia_str, NivelExperiencia.INTERMEDIO)
    
    def _convertir_hora(self, hora_str: str) -> Optional[time]:
        """
        Convierte string a time.
        
        Args:
            hora_str: String con la hora
            
        Returns:
            Optional[time]: Hora convertida o None
        """
        try:
            hora_str = str(hora_str).strip()
            
            # Patrones de hora
            patrones = [
                r'^(\d{1,2}):(\d{2})$',  # HH:MM
                r'^(\d{1,2}):(\d{2}):(\d{2})$',  # HH:MM:SS
                r'^(\d{1,2})$'  # Solo hora
            ]
            
            for patron in patrones:
                match = re.match(patron, hora_str)
                if match:
                    grupos = match.groups()
                    hora = int(grupos[0])
                    minuto = int(grupos[1]) if len(grupos) > 1 else 0
                    segundo = int(grupos[2]) if len(grupos) > 2 else 0
                    
                    return time(hora, minuto, segundo)
            
            return None
            
        except (ValueError, AttributeError):
            return None
    
    def _calcular_estadisticas(self, resultado: ResultadoLectura) -> Dict[str, Any]:
        """
        Calcula estadísticas de la lectura.
        
        Args:
            resultado: Resultado de la lectura
            
        Returns:
            Dict[str, Any]: Estadísticas calculadas
        """
        stats = {
            'total_procesos': len(resultado.procesos),
            'total_recursos': len(resultado.recursos),
            'total_errores': len(resultado.errores),
            'total_advertencias': len(resultado.advertencias),
            'fecha_lectura': datetime.now().isoformat()
        }
        
        # Estadísticas de procesos
        if resultado.procesos:
            tipos_procesos = {}
            prioridades = {}
            
            for proceso in resultado.procesos:
                tipo = proceso.tipo.value
                prioridad = proceso.prioridad.value
                
                tipos_procesos[tipo] = tipos_procesos.get(tipo, 0) + 1
                prioridades[prioridad] = prioridades.get(prioridad, 0) + 1
            
            stats['tipos_procesos'] = tipos_procesos
            stats['prioridades'] = prioridades
            stats['tiempo_total_estimado'] = sum(p.tiempo_estimado_horas for p in resultado.procesos)
        
        # Estadísticas de recursos
        if resultado.recursos:
            tipos_recursos = {}
            capacidad_total = 0
            
            for recurso in resultado.recursos:
                tipo = recurso.tipo.value
                tipos_recursos[tipo] = tipos_recursos.get(tipo, 0) + 1
                capacidad_total += recurso.capacidad_maxima
            
            stats['tipos_recursos'] = tipos_recursos
            stats['capacidad_total'] = capacidad_total
        
        return stats
