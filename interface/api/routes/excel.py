"""
Rutas para Procesamiento de Excel

Este módulo contiene los endpoints para manejar archivos Excel:
- Subir archivos Excel con datos
- Procesar y analizar datos
- Generar reportes en Excel
- Descargar plantillas

Autor: Equipo de Desarrollo
Fecha: 2025-07-07
"""

from typing import Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
import tempfile
import os
from pathlib import Path
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

logger = logging.getLogger(__name__)

# Crear router
router = APIRouter()

@router.post("/procesar", response_model=Dict[str, Any])
async def procesar_archivo_excel(
    archivo: UploadFile = File(..., description="Archivo Excel con procesos y recursos")
):
    """
    Procesa un archivo Excel y devuelve análisis completo con archivo de resultados
    
    Args:
        archivo: Archivo Excel con hojas 'Procesos' y 'Recursos'
        
    Returns:
        Dict con resultados del análisis y ruta del archivo generado
    """
    try:
        logger.info(f"Recibiendo archivo: {archivo.filename}")
        
        # Validar tipo de archivo
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="El archivo debe ser Excel (.xlsx o .xls)"
            )
        
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            # Guardar archivo subido
            content = await archivo.read()
            temp_file.write(content)
            temp_path = temp_file.name
        
        try:
            logger.info(f"Leyendo archivo temporal: {temp_path}")
            
            # Leer datos del Excel
            try:
                df_procesos = pd.read_excel(temp_path, sheet_name="Procesos")
                logger.info(f"Procesos leídos: {len(df_procesos)}")
            except Exception as e:
                logger.error(f"Error leyendo hoja Procesos: {e}")
                raise HTTPException(status_code=400, detail=f"Error leyendo hoja Procesos: {e}")
            
            try:
                df_recursos = pd.read_excel(temp_path, sheet_name="Recursos")
                logger.info(f"Recursos leídos: {len(df_recursos)}")
            except Exception as e:
                logger.error(f"Error leyendo hoja Recursos: {e}")
                raise HTTPException(status_code=400, detail=f"Error leyendo hoja Recursos: {e}")
            
            # Procesar datos
            procesos_count = len(df_procesos)
            recursos_count = len(df_recursos)
            
            # Calcular métricas básicas de manera segura
            tiempo_total = 0
            if 'Tiempo_Estimado_Horas' in df_procesos.columns:
                tiempo_total = float(df_procesos['Tiempo_Estimado_Horas'].sum())
            
            capacidad_total = 0
            if 'Capacidad_Maxima' in df_recursos.columns:
                capacidad_total = float(df_recursos['Capacidad_Maxima'].sum())
            
            costo_promedio = 0
            if 'Costo_Por_Hora' in df_recursos.columns:
                costo_promedio = float(df_recursos['Costo_Por_Hora'].mean())
            
            # Simular análisis
            eficiencia = min(100, (capacidad_total / tiempo_total * 100)) if tiempo_total > 0 else 100
            costo_total = tiempo_total * costo_promedio
            
            logger.info(f"Métricas calculadas - Tiempo: {tiempo_total}, Capacidad: {capacidad_total}, Eficiencia: {eficiencia}")
            
            # Generar archivo de resultados
            try:
                ruta_resultados = generar_excel_resultados(df_procesos, df_recursos, {
                    "tiempo_total": tiempo_total,
                    "capacidad_total": capacidad_total,
                    "eficiencia": eficiencia,
                    "costo_total": costo_total,
                    "procesos_count": procesos_count,
                    "recursos_count": recursos_count
                })
                logger.info(f"Archivo de resultados generado: {ruta_resultados}")
            except Exception as e:
                logger.error(f"Error generando archivo de resultados: {e}")
                raise HTTPException(status_code=500, detail=f"Error generando resultados: {e}")
            
            return {
                "mensaje": "Archivo procesado exitosamente",
                "archivo_original": archivo.filename,
                "procesado": True,
                "procesos_leidos": procesos_count,
                "recursos_leidos": recursos_count,
                "archivo_resultados": Path(ruta_resultados).name,
                "ruta_completa": ruta_resultados,
                "resumen": {
                    "total_procesos": procesos_count,
                    "total_recursos": recursos_count,
                    "eficiencia_proyectada": round(eficiencia, 1),
                    "costo_total": round(costo_total, 2),
                    "tiempo_total_horas": round(tiempo_total, 1),
                    "capacidad_total_horas": round(capacidad_total, 1)
                }
            }
            
        finally:
            # Limpiar archivo temporal
            if os.path.exists(temp_path):
                os.unlink(temp_path)
                logger.info(f"Archivo temporal eliminado: {temp_path}")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error procesando archivo Excel: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando archivo: {str(e)}"
        )


def generar_excel_resultados(df_procesos: pd.DataFrame, df_recursos: pd.DataFrame, metricas: Dict[str, Any]) -> str:
    """Genera un archivo Excel con los resultados del análisis"""
    try:
        # Crear nombre del archivo de salida
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_salida = f"RESULTADOS_PLANIFICACION_{timestamp}.xlsx"
        temp_dir = tempfile.gettempdir()
        ruta_salida = os.path.join(temp_dir, nombre_salida)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # 1. Crear hoja de resumen
        ws_resumen = wb.create_sheet("Resumen_Ejecutivo")
        
        # Título principal
        ws_resumen["A1"] = "REPORTE DE ANÁLISIS DE PLANIFICACIÓN"
        ws_resumen["A1"].font = Font(bold=True, size=16)
        
        # Información general
        ws_resumen["A3"] = "Fecha de Análisis:"
        ws_resumen["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_resumen["A4"] = "Total Procesos:"
        ws_resumen["B4"] = metricas["procesos_count"]
        ws_resumen["A5"] = "Total Recursos:"
        ws_resumen["B5"] = metricas["recursos_count"]
        ws_resumen["A6"] = "Tiempo Total Requerido:"
        ws_resumen["B6"] = f"{metricas['tiempo_total']:.1f} horas"
        ws_resumen["A7"] = "Capacidad Total Disponible:"
        ws_resumen["B7"] = f"{metricas['capacidad_total']:.1f} horas"
        ws_resumen["A8"] = "Eficiencia Estimada:"
        ws_resumen["B8"] = f"{metricas['eficiencia']:.1f}%"
        ws_resumen["A9"] = "Costo Total Estimado:"
        ws_resumen["B9"] = f"${metricas['costo_total']:.2f}"
        
        # 2. Crear hoja de procesos analizados
        ws_procesos = wb.create_sheet("Procesos_Analizados")
        
        # Copiar datos de procesos
        for row in df_procesos.itertuples():
            for col, value in enumerate(row[1:], 1):
                ws_procesos.cell(row=row.Index + 2, column=col, value=value)
        
        # Encabezados
        for col, header in enumerate(df_procesos.columns, 1):
            cell = ws_procesos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # 3. Crear hoja de recursos analizados
        ws_recursos = wb.create_sheet("Recursos_Analizados")
        
        # Copiar datos de recursos
        for row in df_recursos.itertuples():
            for col, value in enumerate(row[1:], 1):
                ws_recursos.cell(row=row.Index + 2, column=col, value=value)
        
        # Encabezados
        for col, header in enumerate(df_recursos.columns, 1):
            cell = ws_recursos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # 4. Crear hoja de recomendaciones
        ws_recom = wb.create_sheet("Recomendaciones")
        ws_recom["A1"] = "RECOMENDACIONES DE OPTIMIZACIÓN"
        ws_recom["A1"].font = Font(bold=True, size=14)
        
        recomendaciones = [
            "Revisar procesos con tiempo estimado muy alto",
            "Balancear la carga entre recursos disponibles",
            "Considerar priorizar procesos críticos",
            "Evaluar la posibilidad de paralelización",
            "Monitorear el cumplimiento de los tiempos estimados"
        ]
        
        for i, rec in enumerate(recomendaciones, 3):
            ws_recom[f"A{i}"] = f"• {rec}"
        
        # Ajustar columnas en todas las hojas
        for ws in wb.worksheets:
            for col in range(1, 10):
                ws.column_dimensions[chr(64 + col)].width = 15
        
        # Guardar archivo
        wb.save(ruta_salida)
        
        logger.info(f"Archivo de resultados generado: {ruta_salida}")
        return ruta_salida
        
    except Exception as e:
        logger.error(f"Error generando Excel de resultados: {str(e)}")
        raise


@router.get("/plantilla")
async def descargar_plantilla():
    """
    Descarga una plantilla Excel para llenar con datos de procesos y recursos
    
    Returns:
        Archivo Excel de plantilla
    """
    try:
        # Crear archivo temporal para la plantilla
        temp_dir = tempfile.gettempdir()
        ruta_plantilla = os.path.join(temp_dir, "plantilla_planificador.xlsx")
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hoja de procesos
        ws_procesos = wb.create_sheet("Procesos")
        headers_procesos = [
            "Nombre", "Descripcion", "Tipo", "Tiempo_Estimado_Horas", 
            "Prioridad", "Recursos_Requeridos"
        ]
        for col, header in enumerate(headers_procesos, 1):
            cell = ws_procesos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Ejemplo de datos
        ws_procesos.cell(row=2, column=1, value="Desarrollo Web")
        ws_procesos.cell(row=2, column=2, value="Crear aplicación web")
        ws_procesos.cell(row=2, column=3, value="rutinario")
        ws_procesos.cell(row=2, column=4, value=40)
        ws_procesos.cell(row=2, column=5, value="alta")
        ws_procesos.cell(row=2, column=6, value="programacion,diseño")
        
        # Crear hoja de recursos
        ws_recursos = wb.create_sheet("Recursos")
        headers_recursos = [
            "Nombre", "Tipo", "Capacidad_Maxima", "Costo_Por_Hora", "Habilidades"
        ]
        for col, header in enumerate(headers_recursos, 1):
            cell = ws_recursos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Ejemplo de datos
        ws_recursos.cell(row=2, column=1, value="Juan Pérez")
        ws_recursos.cell(row=2, column=2, value="humano")
        ws_recursos.cell(row=2, column=3, value=40)
        ws_recursos.cell(row=2, column=4, value=25.50)
        ws_recursos.cell(row=2, column=5, value="programacion,python,javascript")
        
        # Ajustar columnas
        for ws in [ws_procesos, ws_recursos]:
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 15
        
        wb.save(ruta_plantilla)
        logger.info(f"Plantilla Excel creada: {ruta_plantilla}")
        
        return FileResponse(
            path=ruta_plantilla,
            filename="plantilla_planificador.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Error generando plantilla: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error generando plantilla: {str(e)}"
        )


@router.get("/descargar/{nombre_archivo}")
async def descargar_resultados(nombre_archivo: str):
    """
    Descarga un archivo de resultados generado previamente
    
    Args:
        nombre_archivo: Nombre del archivo a descargar
        
    Returns:
        Archivo Excel con resultados
    """
    try:
        # Buscar archivo en directorio temporal
        temp_dir = tempfile.gettempdir()
        ruta_archivo = None
        
        # Buscar archivo que coincida con el patrón
        for archivo in os.listdir(temp_dir):
            if nombre_archivo in archivo and archivo.endswith('.xlsx'):
                ruta_archivo = os.path.join(temp_dir, archivo)
                break
        
        if not ruta_archivo or not os.path.exists(ruta_archivo):
            raise HTTPException(
                status_code=404,
                detail="Archivo no encontrado"
            )
        
        return FileResponse(
            path=ruta_archivo,
            filename=nombre_archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error descargando archivo: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error descargando archivo: {str(e)}"
        )
